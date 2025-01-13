from django.shortcuts import render
from django.contrib.auth import login
from django.contrib.auth.models import User, Group
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.db import models
from courses.utils import get_teachers_with_courses_and_documents
from account.views import user_has_permission
from ders_takip.settings import USER,SUPERUSER,ADMIN
from .xlsxWriter import write_to_excel
from courses.models import Course,CourseFile

@user_has_permission([SUPERUSER,ADMIN])
def erp_view(request):
    distinct_courses = Course.objects.values('name').distinct()
    # GET parametrelerini al
    selected_teacher_id = request.GET.get('teacher_id')
    selected_course_name = request.GET.get('course_name', None)
    teachers_data = get_teachers_with_courses_and_documents(status="aktif", page="erp")
    # Filtreleme işlemi
    if selected_teacher_id:
        selected_teacher_id = int(selected_teacher_id)
        teachers_data = [teacher for teacher in teachers_data if teacher['id'] == selected_teacher_id]
    if selected_course_name:
        for teacher in teachers_data:
            teacher['courses'] = [course for course in teacher['courses'] if course['name'] == selected_course_name]
    return render(request, 'pages/erp.html', {'teachers': teachers_data,
                                                'set_of_courses': distinct_courses,
                                                'selected_teacher_id': int(selected_teacher_id) if selected_teacher_id else None, 
                                                'selected_course_name': selected_course_name})

@user_has_permission([SUPERUSER,ADMIN])
def erp_arsiv_view(request):
    distinct_courses = Course.objects.values('name').distinct()
    # GET parametrelerini al
    selected_teacher_id = request.GET.get('teacher_id')
    selected_course_name = request.GET.get('course_name', None)
    teachers_data = get_teachers_with_courses_and_documents(status="arsiv", page="erp")
    # Filtreleme işlemi
    if selected_teacher_id:
        selected_teacher_id = int(selected_teacher_id)
        teachers_data = [teacher for teacher in teachers_data if teacher['id'] == selected_teacher_id]
    if selected_course_name:
        for teacher in teachers_data:
            teacher['courses'] = [course for course in teacher['courses'] if course['name'] == selected_course_name]
    return render(request, 'pages/erp_arsiv.html', {'teachers': teachers_data,
                                                'set_of_courses': distinct_courses,
                                                'selected_teacher_id': int(selected_teacher_id) if selected_teacher_id else None, 
                                                'selected_course_name': selected_course_name})

@user_has_permission([SUPERUSER,ADMIN])
def erp_iptal_view(request):
    distinct_courses = Course.objects.values('name').distinct()
    # GET parametrelerini al
    selected_teacher_id = request.GET.get('teacher_id')
    selected_course_name = request.GET.get('course_name', None)
    teachers_data = get_teachers_with_courses_and_documents(status="iptal", page="erp")
    # Filtreleme işlemi
    if selected_teacher_id:
        selected_teacher_id = int(selected_teacher_id)
        teachers_data = [teacher for teacher in teachers_data if teacher['id'] == selected_teacher_id]
    if selected_course_name:
        for teacher in teachers_data:
            teacher['courses'] = [course for course in teacher['courses'] if course['name'] == selected_course_name]
    return render(request, 'pages/erp_iptal.html', {'teachers': teachers_data,
                                                'set_of_courses': distinct_courses,
                                                'selected_teacher_id': int(selected_teacher_id) if selected_teacher_id else None, 
                                                'selected_course_name': selected_course_name})


@login_required
def pano_view(request, teacher_id=None):
    distinct_courses = Course.objects.values('name').distinct()
    # GET parametrelerini al
    selected_teacher_id = request.GET.get('teacher_id', teacher_id)
    selected_course_name = request.GET.get('course_name', None)
    teachers_data = get_teachers_with_courses_and_documents(status="aktif", page="pano")
    # Filtreleme işlemi
    if selected_teacher_id:
        selected_teacher_id = int(selected_teacher_id)
        teachers_data = [teacher for teacher in teachers_data if teacher['id'] == selected_teacher_id]
    if selected_course_name:
        for teacher in teachers_data:
            teacher['courses'] = [course for course in teacher['courses'] if course['name'] == selected_course_name]
    return render(request, 'pages/pano.html', {'teachers': teachers_data,
                                                'set_of_courses': distinct_courses,
                                                'selected_teacher_id': int(selected_teacher_id) if selected_teacher_id else None, 
                                                'selected_course_name': selected_course_name})

@login_required
def pano_arsiv_view(request,teacher_id=None):
    distinct_courses = Course.objects.values('name').distinct()
    # GET parametrelerini al
    selected_teacher_id = request.GET.get('teacher_id', teacher_id)
    selected_course_name = request.GET.get('course_name', None)
    teachers_data = get_teachers_with_courses_and_documents(status="arsiv", page="pano")
    # Filtreleme işlemi
    if selected_teacher_id:
        selected_teacher_id = int(selected_teacher_id)
        teachers_data = [teacher for teacher in teachers_data if teacher['id'] == selected_teacher_id]
    if selected_course_name:
        for teacher in teachers_data:
            teacher['courses'] = [course for course in teacher['courses'] if course['name'] == selected_course_name]
    return render(request, 'pages/pano_arsiv.html', {'teachers': teachers_data,
                                                'set_of_courses': distinct_courses,
                                                'selected_teacher_id': int(selected_teacher_id) if selected_teacher_id else None, 
                                                'selected_course_name': selected_course_name})

@login_required
def pano_iptal_view(request,teacher_id=None):
    distinct_courses = Course.objects.values('name').distinct()
    # GET parametrelerini al
    selected_teacher_id = request.GET.get('teacher_id', teacher_id)
    selected_course_name = request.GET.get('course_name', None)
    teachers_data = get_teachers_with_courses_and_documents(status="iptal", page="pano")
    # Filtreleme işlemi
    if selected_teacher_id:
        selected_teacher_id = int(selected_teacher_id)
        teachers_data = [teacher for teacher in teachers_data if teacher['id'] == selected_teacher_id]
    if selected_course_name:
        for teacher in teachers_data:
            teacher['courses'] = [course for course in teacher['courses'] if course['name'] == selected_course_name]
    return render(request, 'pages/pano_iptal.html', {'teachers': teachers_data,
                                                'set_of_courses': distinct_courses,
                                                'selected_teacher_id': int(selected_teacher_id) if selected_teacher_id else None, 
                                                'selected_course_name': selected_course_name})


@login_required
def pano_ozet_view(request):
    data = CourseFile.get_files_in_warning_period()
    return render(request, 'pages/pano_ozet.html', {'data': data})


@user_has_permission([SUPERUSER,ADMIN])
def erp_ozet_view(request):
    data = CourseFile.get_files_in_warning_period()
    return render(request, 'pages/erp_ozet.html', {'data':data})



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.shortcuts import get_object_or_404

@user_has_permission([SUPERUSER, ADMIN])
def excel_view(request, course_id=None, statu=None, page=None):
    """
    Generate and return an Excel file for the given course or all courses.
    """
    # Course bilgisi alınır
    if course_id:
        course = get_object_or_404(Course, id=course_id)
        data = [
            {
                "name": course.teacher.name,
                "surname": course.teacher.surname,
                "title": course.teacher.title,
                "description": course.teacher.description,
                "warning_message": "No issues",
                "courses": [
                    {
                        "name": course.name,
                        "description": course.description,
                        # "warning_message": course.warning_message,
                        "documents": [
                            {
                                # "belge_adi": document.name,
                                # "warning_message": document.warning_message,
                             } #for doc in course.documents.all()
                        ],
                    }
                ]
            }
        ]
        file_name = f"{statu}_{page}_{course.teacher.name}_{course.name}"
    else:
        # Tüm dersler için veri alınır
        data = []
        for course in Course.objects.all():
            teacher = course.teacher
            data.append({
                "name": teacher.name,
                "surname": teacher.surname,
                "title": teacher.title,
                "description": teacher.description,
                "warning_message": "No issues",
                "courses": [
                    {
                        "name": course.name,
                        "description": course.description,
                        "warning_message": course.warning_message,
                        "documents": [
                            {
                                "belge_adi": document.name,
                                "warning_message": doc.warning_message,
                            } for doc in course.documents.all()
                        ],
                    }
                ]
            })
        file_name = f"{statu}_{page}"

    # Excel dosyasını oluştur ve indirilebilir hale getir
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Teachers Overview"

    # Veriyi yaz
    row_index = 1
    column_width = 30
    for teacher in data:
        sheet.cell(row=row_index, column=1, value="Öğretmen Adı")
        sheet.cell(row=row_index, column=2, value=f"{teacher['name']} {teacher['surname']}")
        row_index += 1

        sheet.cell(row=row_index, column=1, value="Unvan")
        sheet.cell(row=row_index, column=2, value=teacher['title'])
        row_index += 1

        sheet.cell(row=row_index, column=1, value="Açıklama")
        sheet.cell(row=row_index, column=2, value=teacher['description'])
        row_index += 1

        sheet.cell(row=row_index, column=1, value="Uyarı Mesaşı")
        sheet.cell(row=row_index, column=2, value=teacher['warning_message'])
        row_index += 2

        for course in teacher['courses']:
            sheet.cell(row=row_index, column=1, value="Ders Adı")
            sheet.cell(row=row_index, column=2, value=course['name'])
            row_index += 1

            sheet.cell(row=row_index, column=1, value="Açıklama")
            sheet.cell(row=row_index, column=2, value=course['description'])
            row_index += 1

            sheet.cell(row=row_index, column=1, value="Warning Message")
            # sheet.cell(row=row_index, column=2, value=str(course['warning_message']))
            row_index += 1

            for document in course['documents']:
                sheet.cell(row=row_index, column=1, value="Belge Adı")
                # sheet.cell(row=row_index, column=2, value=document['belge_adi'])
                sheet.cell(row=row_index, column=3, value=str(document['warning_message']))
                row_index += 1

            row_index += 1
        row_index += 2

    # Sütun genişliklerini ayarla
    for col_index in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_index)
        sheet.column_dimensions[col_letter].width = column_width

    workbook.save(response)
    return response


def excel_view(request,course_id=None, statu=None, page=None):
    if course_id:
        course = get_object_or_404(Course, id=course_id)
        teacher_id = course.teacher.id
        data=get_teachers_with_courses_and_documents(status=statu, page=page, teacher_id=teacher_id, course_id=course_id)
        with open(f"{statu}.txt", "w",encoding="utf-8") as file:
            file.write(str(data))
        write_to_excel(data, f"{statu}_{page}_{course.teacher.name}_{course.name}")
    else:
        data=get_teachers_with_courses_and_documents(status=statu, page=page) 
        write_to_excel(data, f"{statu}_{page}")
    return JsonResponse({"status": "success"})

def index(request):
    user, created = User.objects.get_or_create(
        username="test2_user",
        defaults={
            "email": "test@example.com",
            "is_superuser": True,
            "is_staff": True,
            "password": "your_secure_password_here",  # Güvenli bir parola belirleyin
        }
    )
    if created:
        admin_group, _ = Group.objects.get_or_create(name="Admin")  # Admin grubu yoksa oluştur
        user.groups.add(admin_group)
        user.set_password("your_secure_password_here")  # Şifreyi hashle
        user.save()
    login(request, user)
    return render(request, 'base.html')

