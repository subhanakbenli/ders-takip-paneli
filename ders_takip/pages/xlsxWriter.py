from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def _set_column_width(sheet, column_letters, width):
    """Set the width for given column letters in the Excel sheet."""
    for col_letter in column_letters:
        sheet.column_dimensions[col_letter].width = width

def _write_header(sheet, row_index, headers):
    """Write the header row with bold font."""
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=col_index, value=header)
        cell.font = Font(bold=True)

def _write_row(sheet, row_index, row_data):
    """Write a single row of data."""
    for col_index, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_data)

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.colors import Color

def write_to_excel(teachers_data, file_name):
    """
    Write teacher data with courses and documents to an Excel file with color formatting.
    Creates separate sheets for each teacher with their courses and documents.
    
    Args:
        teachers_data (list): List of teacher data with courses and documents.
        file_name (str): The name of the file (without the .xlsx extension) to save.
    """
    def safe_str(value):
        """Convert any value to string safely, handling None and datetime objects."""
        if value is None:
            return ""
        try:
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d %H:%M:%S')
            return str(value)
        except:
            return ""

    # Define color styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Koyu mavi
    teacher_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")  # Açık mavi
    course_fill = PatternFill(start_color="E4B8B8", end_color="E4B8B8", fill_type="solid")  # Açık kırmızı
    doc_header_fill = PatternFill(start_color="B8E4B8", end_color="B8E4B8", fill_type="solid")  # Açık yeşil
    alternate_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Açık gri

    # Define text styles
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        for teacher in teachers_data:
            try:
                teacher_name = safe_str(teacher.get('name', 'Unknown'))
                teacher_id = safe_str(teacher.get('id', 'NoID'))
                sheet_name = f"{teacher_id}_{teacher_name}"[:31]
                sheet = wb.create_sheet(title=sheet_name)
                
                # Set default alignment for the sheet
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(vertical='center')
                
                # Teacher headers with formatting
                teacher_headers = ["Başlık", "Değer"]
                header_row = sheet.row_dimensions[1]
                header_row.height = 20
                sheet.append(teacher_headers)
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = white_font
                
                # Teacher information with formatting
                teacher_info = [
                    ["Öğretmen Adı", safe_str(teacher.get("name"))],
                    ["Soyadı", safe_str(teacher.get("surname"))],
                    ["Ünvan", safe_str(teacher.get("title"))],
                    ["Açıklama", safe_str(teacher.get("description"))],
                    ["Telefon", safe_str(teacher.get("telephone"))],
                    ["Telefon 2", safe_str(teacher.get("telephone2"))],
                    ["E-posta", safe_str(teacher.get("mail"))],
                    ["Adres", safe_str(teacher.get("adress"))],
                    ["Oluşturulma Tarihi", safe_str(teacher.get("created_at"))],
                    ["Oluşturan", safe_str(teacher.get("created_by"))]
                ]
                
                for idx, info in enumerate(teacher_info, 2):
                    sheet.append(info)
                    for cell in sheet[idx]:
                        cell.fill = teacher_fill
                        if cell.column == 1:  # First column
                            cell.font = bold_font
                
                sheet.append([])  # Empty row
                
                for course in teacher.get("courses", []):
                    try:
                        # Course header
                        current_row = sheet.max_row + 1
                        sheet.append(["DERS BİLGİLERİ", ""])
                        for cell in sheet[current_row]:
                            cell.fill = header_fill
                            cell.font = white_font
                        
                        # Course information
                        course_info = [
                            ["Ders Adı", safe_str(course.get("name"))],
                            ["Pano Durumu", safe_str(course.get("statu_pano"))],
                            ["ERP Durumu", safe_str(course.get("statu_erp"))],
                            ["Açıklama", safe_str(course.get("description"))],
                            ["Başlangıç Tarihi", safe_str(course.get("start_date"))],
                            ["Bitiş Tarihi", safe_str(course.get("end_date"))],
                            ["Dilekçe Gerekli", "Evet" if course.get("dilekce_required") else "Hayır"],
                            ["Oluşturulma Tarihi", safe_str(course.get("created_at"))],
                            ["Oluşturan", safe_str(course.get("created_by"))]
                        ]
                        
                        for info in course_info:
                            current_row = sheet.max_row + 1
                            sheet.append(info)
                            for cell in sheet[current_row]:
                                cell.fill = course_fill
                                if cell.column == 1:
                                    cell.font = bold_font
                        
                        sheet.append([])
                        
                        # Documents section
                        documents = course.get("documents", [])
                        if documents:
                            doc_headers = [
                                "Belge ID", "Kategori", "Belge Adı", "Yüklendi mi?", "Yüklenme Tarihi",
                                "Dilekçe Adı", "Dilekçe Yüklendi mi?", "Başlangıç Tarihi", "Bitiş Tarihi",
                                "Tip", "Pano Durumu", "ERP Durumu", "Etkinlik No", "Etkinlik Adı",
                                "Kodu", "Etkinlik Tarihi", "Etkinlik Açıklaması", "Öğretmen Adı",
                                "Sınıf", "Şehir", "Katılımcılar", "Katılımcı Sayısı"
                            ]
                            
                            current_row = sheet.max_row + 1
                            sheet.append(doc_headers)
                            for cell in sheet[current_row]:
                                cell.fill = doc_header_fill
                                cell.font = bold_font
                            
                            # Document rows with alternating colors
                            for idx, doc in enumerate(documents, 1):
                                try:
                                    doc_row = [
                                        safe_str(doc.get("id")),
                                        safe_str(doc.get("category")),
                                        safe_str(doc.get("belge_adi")),
                                        "Evet" if doc.get("is_uploaded") else "Hayır",
                                        safe_str(doc.get("uploaded_at")),
                                        safe_str(doc.get("dilekce_name")),
                                        "Evet" if doc.get("dilekce_is_uploaded") else "Hayır",
                                        safe_str(doc.get("start_date")),
                                        safe_str(doc.get("end_date")),
                                        safe_str(doc.get("type")),
                                        safe_str(doc.get("statu_pano")),
                                        safe_str(doc.get("statu_erp")),
                                        safe_str(doc.get("etkinlik_no")),
                                        safe_str(doc.get("etkinlik_adi")),
                                        safe_str(doc.get("kodu")),
                                        safe_str(doc.get("etkinklik_tarihi")),
                                        safe_str(doc.get("etkinlik_aciklamasi")),
                                        safe_str(doc.get("ogretmen_adi")),
                                        safe_str(doc.get("sinif")),
                                        safe_str(doc.get("sehir")),
                                        safe_str(doc.get("katilanlar")),
                                        safe_str(doc.get("katilimci_sayisi"))
                                    ]
                                    current_row = sheet.max_row + 1
                                    sheet.append(doc_row)
                                    
                                    # Alternate row colors
                                    if idx % 2 == 0:
                                        for cell in sheet[current_row]:
                                            cell.fill = alternate_row_fill
                                            
                                except Exception as e:
                                    print(f"Error processing document in course {safe_str(course.get('name', 'Unknown'))}: {str(e)}")
                                    continue
                            
                            sheet.append([])  # Empty row after documents
                    
                    except Exception as e:
                        print(f"Error processing course for teacher {teacher_name}: {str(e)}")
                        continue
                
                # Adjust column widths
                try:
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        
                        adjusted_width = min((max_length + 2), 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
                
                except Exception as e:
                    print(f"Error adjusting column widths for sheet {sheet_name}: {str(e)}")
            
            except Exception as e:
                print(f"Error processing teacher {safe_str(teacher.get('name', 'Unknown'))}: {str(e)}")
                continue
        
        try:
            wb.save(f"{file_name}.xlsx")
            print(f"Data successfully written to {file_name}.xlsx")
        except Exception as e:
            print(f"Error saving workbook: {str(e)}")
            raise
    
    except Exception as e:
        print(f"Critical error in write_to_excel: {str(e)}")
        raise